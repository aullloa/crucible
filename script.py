# Alfredo Ulloa
# COMP 467
# Project 4
# The Crucible

import argparse
import sys
from pymongo import MongoClient, errors
import ffmpeg
import vimeo
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
import csv

# Argparse functions
parser = argparse.ArgumentParser()
parser.add_argument("--baselight", "-bl", action="store", help="import baselight file into db")
parser.add_argument("--xytech", "-xt",action="store", help="import xytech file into db")
parser.add_argument("--process", action="store", help="process video file")
parser.add_argument("--output", action="store_true", help="export data into XLS file")

args = parser.parse_args()

# Set up vimeo client
vimeo_client = vimeo.VimeoClient(
    token = "0070400c399f5e91c87ad0e80960999c",
    client_id='6b1666c5d4c95d3ecd5941b8dec63a7dab93112d',
    client_secret='xUzuhK2OYRA2ogYXxZ8eAqouKylGckReL4LJyll4Wg/oXJ4b18bFQIZBY1y7uwjIm+SlMrQgm4VbG/Dg0So3NVV/Z4q0ESksXeFB/42hkH3D9foA9m/7TwAFv72rWsPg',
)

# Connect to DB
client = MongoClient('localhost', 27017)
database = client["crucible"]

# Helper methods
def timecode(frame, fps):
    frames = frame % fps
    total_seconds  = frame // fps
    seconds = total_seconds % 60
    total_minutes = total_seconds // 60
    minutes = total_minutes % 60
    total_hours = total_minutes // 60
    hours = total_hours % 24
    return f"{hours:02}:{minutes:02}:{seconds:02}:{frames:02}"

def find_new_url(baselight_folder, xytech):
    for url in xytech.find():
        if "Planeshifter" in baselight_folder and baselight_folder in url["location"]:
            return url["Workorder"], url["location"]
    return None, None

def find_timecode_range(frame):
    pre_frame = frame - 48
    post_frame = frame + 48
    start = timecode(pre_frame, fps)
    end = timecode(post_frame, fps)
    return f"{start}-{end}"

def get_vimeo_data():
    # Export data from Vimeo account to a csv
    csv_path = "output/vimeo_data.csv"
    with open(csv_path, "w", newline="") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(["Title", "URI", "Public Link", "Status"])
        response = vimeo_client.get("/me/videos")

        while True:
            data = response.json()
            for video in data["data"]:
                title = video.get("name")
                uri = video.get("uri")
                link = video.get("link")
                status = video.get("status")
                writer.writerow([title, uri, link, status])

            next_page = data.get("paging").get("next")
            if not next_page:
                break
            response = vimeo_client.get(next_page)

    print("Vimeo data exported to csv!")

def create_thumbnail(frame):
    timestamp = frame / fps
    output_path = os.path.join(output_folder, f"{frame}_thumbnail.png")

    (
        ffmpeg
        .input(args.process, ss=timestamp)
        .output(output_path, vframes=1, s="96x74")
        .overwrite_output()
        .run()
    )

    return output_path

def create_clip(frame):
    # Create clip and upload to Vimeo
    output_video_path = os.path.join(output_folder, f"{frame}.mp4")
    start = (frame - 48) / fps
    end = (frame + 48) / fps
    duration = end - start
    timecode_range = find_timecode_range(frame)

    (
        ffmpeg
        .input(args.process, ss=start, t=duration)
        .output(output_video_path)
        .overwrite_output()
        .run()
    )

    video = vimeo_client.upload(
        output_video_path,
        data={
            "name": f"Clip for {timecode_range}"
        }
    )
    print(f"Clip {timecode_range} created and uploaded successfully!")

# Import baselight data to DB
# Excludes preliminary baselight1 directory
if args.baselight:
    collection_name = "baselight"
    collection = database[collection_name]
    try:
        baselight_file = open(args.baselight, "r")
        baselight_lines = baselight_file.readlines()

        for line in baselight_lines:
            each_line = line.strip().split()
            url = each_line[0].strip("/baselightfilesystem1")
            frames = list(map(int, each_line[1:]))
            collection.insert_one({"folder": url, "frames": frames})
        print(f"Data from {args.baselight} imported successfully!")
    except FileNotFoundError:
        print(f"{args.baselight} file not found")
        sys.exit(1)
    except errors.ConnectionFailure:
        print("Could not connect to MongoDB")
        sys.exit(1)

# Import xytech data to DB
if args.xytech:
    collection_name = "xytech"
    collection = database[collection_name]
    try:
        xytech_file = open(args.xytech, "r")
        xytech_lines = xytech_file.readlines()
        for line in xytech_lines:
            if line.__contains__("Workorder"):
                workorder_line = line.strip().split()
                workorder = workorder_line[2]
                continue
            each_line = line.strip().split()
            if not (line.startswith("/")):
                continue
            else:
                url = each_line[0].strip()
                collection.insert_one({"Workorder": workorder, "location": url})
        print(f"Data from {args.xytech} imported successfully!")
    except FileNotFoundError:
        print(f"{args.xytech} file not found")
        sys.exit(1)
    except errors.ConnectionFailure:
        print("Could not connect to MongoDB")
        sys.exit(1)

if args.process:
    baselight_db = database["baselight"]
    xytech_db = database["xytech"]
    results = []
    fps = 24

    # Extract timecode from video
    video_data = ffmpeg.probe(args.process)
    for i in video_data["streams"]:
        if i["index"] == 0:
            video_timecode = i["tags"]["timecode"]
            print(f"Timecode extracted from {args.process} is {video_timecode}")
            print("*" * 40)

    # Find all data corresponding to video
    contents = baselight_db.find()
    for content in contents:
        folder = content["folder"]
        if "Planeshifter"  not in folder:
            continue

        workorder, location = find_new_url(folder, xytech_db)
        if workorder is None or location is None:
            continue

        for frame in content["frames"]:
            results.append({
                "workorder": workorder,
                "location": location,
                "frame": frame,
                "frame range": f"{frame - 48} - {frame + 48}",
                "timecode range": find_timecode_range(frame),
            })

if args.output:
    rows = []
    output_folder = "output"

    for result in results:
        frame = result["frame"]
        workorder = result["workorder"]
        location = result["location"]
        timecode_range = result["timecode range"]

        thumbnail = create_thumbnail(frame)

        rows.append({
            "Workorder": workorder,
            "Location": location,
            "Frame Range": f"{frame - 48} - {frame + 48}",
            "Timecode Range": timecode_range,
            "Thumbnail": thumbnail,
        })

        create_clip(frame)

    df = pd.DataFrame(rows)
    df.to_excel("output/output.xlsx", index=False)

    # Adding the thumbnails to excel
    wb = load_workbook("output/output.xlsx")
    ws = wb.active

    thumbnails = df.columns.get_loc("Thumbnail") + 1

    for row in range(2, len(df) + 2):
        img_path = ws.cell(row=row, column=thumbnails).value
        if img_path and os.path.exists(img_path):
            img = Image(img_path)
            img.width = 96
            img.height = 74
            ws.add_image(img, ws.cell(row=row, column=thumbnails).coordinate)
            ws.cell(row=row, column=thumbnails).value = ""

    wb.save("output/output.xlsx")
    print(f"Data exported output/output.xlsx successfully!")

    get_vimeo_data()

